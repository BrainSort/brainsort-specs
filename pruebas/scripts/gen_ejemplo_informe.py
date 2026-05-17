"""Genera 3.3 Informe de Prueba EJEMPLO llenado con resultados simulados de BrainSort"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = Workbook()
hf = Font(name='Calibri',bold=True,color='FFFFFF',size=11)
hfill = PatternFill(start_color='1F3A5F',end_color='1F3A5F',fill_type='solid')
green = PatternFill(start_color='D4EDDA',end_color='D4EDDA',fill_type='solid')
red = PatternFill(start_color='FFD7D7',end_color='FFD7D7',fill_type='solid')
yellow = PatternFill(start_color='FFF3CD',end_color='FFF3CD',fill_type='solid')
gray = PatternFill(start_color='E9ECEF',end_color='E9ECEF',fill_type='solid')
bdr = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
wrap = Alignment(wrap_text=True,vertical='top')

def do_hdr(ws, row=1):
    for c in ws[row]:
        c.font=hf; c.fill=hfill; c.border=bdr
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

def do_row(ws, row, fill=None):
    for c in ws[row]:
        c.border=bdr; c.alignment=wrap
        if fill: c.fill=fill

# ===== HOJA 1: RESUMEN EJECUTIVO =====
ws = wb.active; ws.title = 'Resumen Ejecutivo'
ws.append(['INFORME DE PRUEBA DE SOFTWARE - BrainSort Sprint 3'])
ws.merge_cells('A1:D1')
ws['A1'].font = Font(bold=True,size=16,color='1F3A5F')
ws.append(['Fecha de Ejecucion: 31/05/2026  |  Ciclo: 1  |  Modulo: Simulacion Visual'])
ws.merge_cells('A2:D2')
ws['A2'].font = Font(size=11,italic=True)
ws.append([])
ws.append(['Metrica','Valor']); do_hdr(ws,4)
metrics = [
    ['Total de Casos de Prueba','19'],
    ['Ejecutados','19 / 19'],
    ['Pasaron','16'],
    ['Fallaron','2'],
    ['Bloqueados','1'],
    ['No ejecutados','0'],
    ['Tasa de exito','84.2%'],
    ['Defectos encontrados','3'],
    ['Defectos Criticos abiertos','1'],
]
for m in metrics: ws.append(m); do_row(ws,ws.max_row)
ws.append([])
ws.append(['VEREDICTO: RECHAZADO - 1 defecto critico abierto (DEF-001: timeout no implementado)'])
ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
ws[f'A{ws.max_row}'].font = Font(bold=True,color='CC0000',size=12)
ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 25

# ===== HOJA 2: RESULTADOS DETALLADOS =====
ws2 = wb.create_sheet('Resultados Detallados')
ws2.append(['ID','Modulo','Descripcion','Severidad','Resultado','Observaciones'])
do_hdr(ws2)

results = [
['CP-SIM-01','Simulacion','Play inicia animacion','Critica','Paso','Animacion inicia en <100ms. isPlaying=true verificado en React DevTools.'],
['CP-SIM-02','Simulacion','Pausa detiene animacion','Critica','Paso','Se congela en paso exacto. currentStep no cambia despues de pause.'],
['CP-SIM-03','Simulacion','Color coding Azul/Amarillo/Rojo/Verde','Critica','Paso','Colores verificados: Azul=#2196F3, Amarillo=#FFC107, Rojo=#DC3545, Verde=#28A745. Coinciden con constitution.md.'],
['CP-SIM-04','Simulacion','Velocidad 0.25x a 2.0x','Alta','Paso','8 niveles funcionan. A 0.25x cada paso ~4s, a 2.0x ~0.5s. Clamp en setSpeed() funciona correctamente.'],
['CP-SIM-05','Simulacion','Pseudocodigo sincronizado','Alta','Fallo','DEF-002: En Insertion Sort, lineaPseudocodigo se desfasa 1 paso despues del step 8. Bubble y Selection OK.'],
['CP-SIM-06','Simulacion','Animacion completa sin detenerse','Critica','Paso','Probado con 8, 10, 12 y 15 elementos. isCompleted=true al final. isPlaying=false automatico.'],
['CP-SIM-07','Simulacion','Controles al finalizar','Alta','Paso','togglePlayPause() no hace nada si isCompleted=true. resetSimulation() funciona correctamente.'],
['CP-SIM-08','Simulacion','Mensaje completado + opciones','Alta','Paso','CompletionOverlay aparece con 3 opciones. Desaparece a los 5.1s (aceptable). No bloquea navegacion.'],
['CP-SIM-09','Simulacion','Timeout bucle infinito','Critica','Fallo','DEF-001 CRITICO: No hay timeout. Con datos corruptos el do-while en generateBubbleSortSteps nunca termina. App se congela.'],
['CP-SIM-10','Simulacion','Feedback daltonicos','Media','Bloqueado','DEF-003: Iconos de accesibilidad no implementados en Bar.tsx. Solo colores. Feature pendiente.'],
['CP-DAT-01','Datos','Datos predeterminados 8-15 elem','Critica','Paso','Genero 12 elementos primera carga, 9 segunda. Nunca ordenados. data[] verificado.'],
['CP-DAT-02','Datos','Generar nuevos datos','Alta','Paso','Arreglo cambia cada vez. Probado 10 veces sin repeticion. steps[] se recalculan.'],
['CP-DAT-03','Datos','Datos personalizados validos','Alta','Paso','[5,2,8,1,9,3] -> 6 barras correctas. data[] refleja input exacto.'],
['CP-DAT-04','Datos','Datos invalidos muestran error','Alta','Paso','Variantes A,B,C,D probadas. Sanitizacion funciona. Simulacion no inicia con datos invalidos.'],
['CP-CTL-01','Controles','Reiniciar simulacion','Alta','Paso','resetSimulation() ejecuta correctamente: currentStep=0, isPlaying=false, isCompleted=false.'],
['CP-ENG-01','Engine','Bubble Sort pasos correctos','Critica','Paso','generateBubbleSortSteps([5,2,8,1]) genera 14 steps. Ultimo: estadoArray=[1,2,5,8]. Estructura correcta.'],
['CP-ENG-02','Engine','Selection Sort pasos correctos','Critica','Paso','Estructura identica a Bubble. Array final ordenado [1,2,5,8].'],
['CP-ENG-03','Engine','Insertion Sort pasos correctos','Critica','Paso','Array final ordenado. NOTA: lineaPseudocodigo en step 8 parece incorrecto (ver DEF-002).'],
['CP-REN-01','Rendimiento','>=24 FPS simulacion','Critica','Paso','Bubble 15 elem Pixel 5a: 28 FPS promedio, 22 FPS minimo (pico breve). Selection: 31 FPS. Insertion: 30 FPS.'],
]

fills = {'Paso':green,'Fallo':red,'Bloqueado':yellow}
for r in results:
    ws2.append(r); do_row(ws2,ws2.max_row,fills.get(r[4]))

widths2 = [12,14,35,12,12,60]
for i,w in enumerate(widths2): ws2.column_dimensions[chr(65+i)].width = w
ws2.auto_filter.ref = ws2.dimensions; ws2.freeze_panes = 'A2'

# ===== HOJA 3: CRITERIOS =====
ws3 = wb.create_sheet('Criterios Aceptacion')
ws3.append(['Criterio','Umbral','Valor Obtenido','Cumple']); do_hdr(ws3)
crit_data = [
    ['Cobertura engine (src/engine/)','>=85%','87.3%','Si'],
    ['Casos criticos pasados (6/6)','100%','83.3% (5/6 - CP-SIM-09 fallo)','No'],
    ['Casos totales pasados','>=95%','84.2% (16/19)','No'],
    ['Defectos bloqueantes abiertos','0','1 (DEF-001)','No'],
    ['FPS simulacion (Pixel 5a)','>=24 FPS','28 FPS promedio','Si'],
    ['Carga SimulationScreen','<3 segundos','1.8 segundos','Si'],
]
for c in crit_data:
    ws3.append(c)
    f = green if c[3]=='Si' else red
    do_row(ws3,ws3.max_row,f)
for i,w in enumerate([38,25,35,12]): ws3.column_dimensions[chr(65+i)].width = w

# ===== HOJA 4: DEFECTOS =====
ws4 = wb.create_sheet('Registro de Defectos')
ws4.append(['ID','Severidad','Modulo','Descripcion','CP','Estado','Asignado','Fecha','Resolucion'])
do_hdr(ws4)

defectos = [
['DEF-001','Critica','Engine / SimulationContext',
 'No existe timeout de seguridad contra bucle infinito. El do-while en generateBubbleSortSteps (mock-bubble-sort.ts linea 33) puede nunca terminar si el array tiene datos que causan swapped=true infinitamente. La app se congela sin forma de recuperarse. Solucion propuesta: agregar contador MAX_STEPS=10000 en SimulationContext o en el engine.',
 'CP-SIM-09','Abierto','Carlos Mendoza','28/05/2026','Pendiente - Prioridad Sprint 4'],

['DEF-002','Alta','Engine (Insertion Sort)',
 'En mock-insertion-sort.ts, lineaPseudocodigo del step 8 apunta a linea 2 pero deberia ser linea 3. Causa: el indice de pseudocodigo no se actualiza correctamente cuando el elemento se inserta en la primera posicion. Fix: corregir linea 45 de mock-insertion-sort.ts.',
 'CP-SIM-05','Resuelto','Ana Rodriguez','29/05/2026','Fix en commit abc123 el 29/05. Re-test pendiente.'],

['DEF-003','Media','UI (Bar.tsx)',
 'Los iconos de feedback para daltonicos (check, X) no estan implementados. El componente Bar.tsx solo renderiza rectangulos de color sin iconos de estado. Segun HU-06 y constitution.md, se requieren iconos ademas de colores para accesibilidad.',
 'CP-SIM-10','Abierto','Sin asignar','30/05/2026','Feature pendiente. No es bloqueante para release.'],
]

sev_fills = {'Critica':red,'Alta':yellow,'Media':PatternFill(start_color='D1ECF1',end_color='D1ECF1',fill_type='solid')}
for d in defectos:
    ws4.append(d); do_row(ws4,ws4.max_row,sev_fills.get(d[1]))
widths4 = [10,10,20,55,12,12,16,12,40]
for i,w in enumerate(widths4): ws4.column_dimensions[chr(65+i)].width = w

# ===== HOJA 5: RENDIMIENTO =====
ws5 = wb.create_sheet('Rendimiento')
ws5.append(['Algoritmo','Elementos','Dispositivo','FPS Promedio','FPS Minimo','Cumple >=24']); do_hdr(ws5)
perf = [
    ['Bubble Sort','8','Pixel 5a (Android 14)','42','38','Si'],
    ['Bubble Sort','15','Pixel 5a (Android 14)','28','22','Si (min bajo pero aceptable)'],
    ['Selection Sort','8','Pixel 5a (Android 14)','45','40','Si'],
    ['Selection Sort','15','Pixel 5a (Android 14)','31','26','Si'],
    ['Insertion Sort','8','Pixel 5a (Android 14)','44','39','Si'],
    ['Insertion Sort','15','Pixel 5a (Android 14)','30','24','Si (justo en el limite)'],
    ['Bubble Sort','15','Chrome 126 (Windows 11)','60','58','Si'],
    ['Bubble Sort','15','Chrome 126 (Web, throttled 4x)','35','28','Si'],
]
for p in perf:
    f = green if 'Si' in p[5] else red
    ws5.append(p); do_row(ws5,ws5.max_row,f)
for i,w in enumerate([18,12,28,14,14,28]): ws5.column_dimensions[chr(65+i)].width = w

# ===== HOJA 6: CONCLUSIONES =====
ws6 = wb.create_sheet('Conclusiones')
ws6.append(['CONCLUSIONES Y RECOMENDACIONES']); ws6['A1'].font = Font(bold=True,size=14,color='1F3A5F')
ws6.append([])
ws6.append(['CONCLUSIONES:']); ws6[f'A{ws6.max_row}'].font = Font(bold=True,size=12)
concl = [
    '1. El modulo de Simulacion Visual es funcionalmente estable para Bubble Sort y Selection Sort.',
    '2. DEF-001 (Critico): La ausencia de timeout en el engine es un riesgo de produccion. Si datos corruptos causan ciclo infinito, la app se congela.',
    '3. DEF-002 (Resuelto): El desfase de pseudocodigo en Insertion Sort fue corregido el 29/05. Pendiente re-test.',
    '4. El rendimiento cumple >=24 FPS, pero Insertion Sort con 15 elementos esta al limite (24 FPS minimo).',
    '5. Cobertura de engine al 87.3% supera el umbral de 85%.',
]
for c in concl: ws6.append([c])
ws6.append([])
ws6.append(['RECOMENDACIONES:']); ws6[f'A{ws6.max_row}'].font = Font(bold=True,size=12)
recs = [
    '1. URGENTE: Implementar timeout en SimulationContext.tsx - max 10,000 pasos o 30s de ejecucion. Desbloquea CP-SIM-09.',
    '2. Re-testear CP-SIM-05 despues del fix de DEF-002 (commit abc123).',
    '3. Implementar iconos de accesibilidad en Bar.tsx para daltonicos (DEF-003).',
    '4. Optimizar InsertionSortAnimation.tsx para mejorar FPS minimo con 15 elementos.',
    '5. Agregar pruebas unitarias para edge cases: arreglo de 1 elemento, arreglo con duplicados.',
]
for r in recs: ws6.append([r])
ws6.append([])
ws6.append(['ACCIONES PENDIENTES:']); ws6[f'A{ws6.max_row}'].font = Font(bold=True,size=12)
ws6.append(['Accion','Responsable','Fecha Compromiso','Estado']); do_hdr(ws6,ws6.max_row)
actions = [
    ['Implementar timeout en SimulationContext','Carlos Mendoza','02/06/2026','En Progreso'],
    ['Re-test CP-SIM-05 (fix DEF-002)','Luis Perez','01/06/2026','Pendiente'],
    ['Iconos accesibilidad Bar.tsx','Ana Rodriguez','05/06/2026','Pendiente'],
    ['Ciclo de pruebas #2 (re-test)','Luis Perez','03/06/2026','Pendiente'],
]
for a in actions: ws6.append(a); do_row(ws6,ws6.max_row)
ws6.column_dimensions['A'].width = 55

# ===== HOJA 7: APROBACIONES =====
ws7 = wb.create_sheet('Aprobaciones')
ws7.append(['Nombre','Rol','Firma','Fecha']); do_hdr(ws7)
aprs = [
    ['Luis Perez Garcia','QA Lead','[firmado]','31/05/2026'],
    ['Carlos Mendoza','Tech Lead','[firmado - con nota: timeout prioridad Sprint 4]','31/05/2026'],
    ['Dra. Maria Lopez','Product Owner','[rechazado - requiere fix DEF-001]','31/05/2026'],
]
for a in aprs: ws7.append(a); do_row(ws7,ws7.max_row)
for i,w in enumerate([25,18,45,15]): ws7.column_dimensions[chr(65+i)].width = w

out = os.path.join(os.path.dirname(os.path.abspath(__file__)),'..','ejemplos','3.3-Informe-de-Prueba-EJEMPLO.xlsx')
wb.save(out)
print('OK: '+out)
