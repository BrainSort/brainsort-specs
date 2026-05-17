import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_report(json_path, output_path):
    if not os.path.exists(json_path):
        print(f"Error: No se encontro el archivo {json_path}")
        sys.exit(1)

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ejecución (Jest)"

    # Estilos
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
    green = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    red = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap = Alignment(wrap_text=True, vertical='center')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Cabecera
    ws.append(['REPORTE AUTOMATIZADO DE PRUEBAS - BrainSort (React Native)'])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=16, color='1F3A5F')
    
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([f'Fecha de Ejecución: {date_str}'])
    ws.merge_cells('A2:E2')
    ws.append([])

    # Resumen
    ws.append(['Métrica', 'Valor'])
    for cell in ws[4]:
        cell.font = hf
        cell.fill = hfill
        cell.border = bdr

    ws.append(['Total de Pruebas', data.get('numTotalTests', 0)])
    ws.append(['Pruebas que Pasaron', data.get('numPassedTests', 0)])
    ws.append(['Pruebas que Fallaron', data.get('numFailedTests', 0)])
    for r in range(5, 8):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = bdr

    ws.append([])
    
    # Detalle de Casos
    headers = ['Suite / Archivo', 'Caso de Prueba', 'Estado', 'Duración (ms)', 'Errores']
    ws.append(headers)
    hdr_row = ws.max_row
    for idx, c in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=idx)
        cell.font = hf
        cell.fill = hfill
        cell.border = bdr
        cell.alignment = center

    # Parsear resultados
    for test_suite in data.get('testResults', []):
        suite_name = os.path.basename(test_suite.get('name', ''))
        
        for assertion in test_suite.get('assertionResults', []):
            title = assertion.get('title', '')
            status = assertion.get('status', 'unknown')
            duration = assertion.get('duration', 0)
            failure_msgs = "\n".join(assertion.get('failureMessages', []))
            
            row = [suite_name, title, status.upper(), duration, failure_msgs]
            ws.append(row)
            
            curr_row = ws.max_row
            for col_idx in range(1, 6):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.border = bdr
                cell.alignment = wrap
            
            # Color estado
            status_cell = ws.cell(row=curr_row, column=3)
            if status == 'passed':
                status_cell.fill = green
            elif status == 'failed':
                status_cell.fill = red

    # Ajustar columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Reporte generado exitosamente en: {output_path}")

if __name__ == "__main__":
    app_dir = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'brainsort-app')
    json_result = os.path.join(app_dir, 'test-results.json')
    out_excel = os.path.join(os.path.dirname(__file__), '..', 'ejemplos', 'Reporte_Automatizado_Jest.xlsx')
    
    generate_report(json_result, out_excel)
