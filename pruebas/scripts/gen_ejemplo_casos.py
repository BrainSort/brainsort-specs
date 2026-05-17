"""Genera 3.2 Casos de Prueba EJEMPLO llenado con datos reales de BrainSort"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = Workbook()
hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
hfill = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
crit = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
alta = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
med = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
bdr = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
wrap = Alignment(wrap_text=True, vertical='top')

def do_hdr(ws, row=1):
    for c in ws[row]:
        c.font=hf; c.fill=hfill; c.border=bdr
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

def do_row(ws, row, sev=None):
    for c in ws[row]:
        c.border=bdr; c.alignment=wrap
        if sev=='Critica': c.fill=crit
        elif sev=='Alta': c.fill=alta
        elif sev=='Media': c.fill=med

# HOJA 1: CASOS
ws = wb.active; ws.title = 'Casos de Prueba'
ws.append(['ID','Modulo','Nombre','HU/CO','Precondiciones','Datos de Entrada','Pasos','Resultado Esperado','Severidad','Estado'])
do_hdr(ws)

cases = [
['CP-SIM-01','Simulacion','Play inicia animacion','HU-04',
 '1. Usuario autenticado como Estudiante\n2. Bubble Sort seleccionado\n3. SimulationScreen cargada con [5,2,8,1]\n4. Estado: Pausa (isPlaying=false en SimulationContext)',
 'Tap en boton Play (ControlBar.tsx)',
 '1. Verificar que ControlBar muestra icono "play-arrow"\n2. Presionar boton Play\n3. Observar que la animacion comienza\n4. Verificar que ControlBar cambia a icono "pause"',
 'Barras se animan. isPlaying=true en SimulationContext. Icono cambia a "pause". currentStep incrementa automaticamente.','Critica','Pendiente'],

['CP-SIM-02','Simulacion','Pausa detiene animacion','HU-04',
 '1. Simulacion en ejecucion (isPlaying=true)\n2. currentStep > 0',
 'Tap en boton Pausa (ControlBar.tsx)',
 '1. Con animacion corriendo, presionar Pausa\n2. Observar que la animacion se detiene\n3. Verificar currentStep no cambia',
 'Animacion congelada en paso actual. isPlaying=false. Icono vuelve a "play-arrow".','Critica','Pendiente'],

['CP-SIM-03','Simulacion','Color coding: Azul/Amarillo/Rojo/Verde','HU-04',
 '1. Bubble Sort con arreglo [5,2,8,1]\n2. Velocidad 0.5x para observar',
 'Arreglo [5,2,8,1], Bubble Sort, velocidad 0.5x',
 '1. Play. Paso 1 (comparar 5 vs 2): verificar AMARILLO en indices 0,1\n2. Paso 2 (intercambiar): verificar ROJO en indices 0,1\n3. Elementos no involucrados: verificar AZUL\n4. Al final primera pasada: ultimo elemento VERDE\n5. Al completar: TODOS verde',
 'Azul=#2196F3 (inactivo), Amarillo=#FFC107 (comparacion, tipoOperacion="comparison"), Rojo=#DC3545 (intercambio, tipoOperacion="swap"), Verde=#28A745 (posicion final)','Critica','Pendiente'],

['CP-SIM-04','Simulacion','Velocidad ajustable 0.25x a 2.0x','HU-04',
 '1. SimulationScreen cargada\n2. SpeedSlider.tsx visible',
 'Valores: 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2.0',
 '1. Para cada valor: mover SpeedSlider\n2. Verificar que setSpeed() en SimulationContext recibe el valor\n3. Verificar clamp: Math.max(0.25, Math.min(2.0, newSpeed))\n4. Play y observar cambio de ritmo visual',
 '8 niveles. A 0.25x cada paso ~4s, a 2.0x ~0.5s. speed en PlaybackState refleja valor correcto.','Alta','Pendiente'],

['CP-SIM-05','Simulacion','Pseudocodigo sincronizado con paso','HU-04',
 '1. PseudocodePanel.tsx visible\n2. Bubble Sort con [5,2,8,1]',
 'Bubble Sort, BUBBLE_SORT_PSEUDOCODE tiene 12 lineas',
 '1. Play. En paso de comparacion: verificar lineaPseudocodigo=6 resaltada\n2. En paso de intercambio: verificar lineaPseudocodigo=7\n3. En paso idle inicial: verificar lineaPseudocodigo=2\n4. Al final: lineaPseudocodigo=12',
 'PseudocodePanel resalta la linea correcta segun steps[currentStep].lineaPseudocodigo','Alta','Pendiente'],

['CP-SIM-06','Simulacion','Animacion completa sin detenerse','HU-06',
 '1. Bubble Sort con 10 elementos aleatorios\n2. Velocidad 1.0x',
 'Arreglo de 10 elementos, sin interaccion del usuario',
 '1. Play y NO interactuar\n2. Esperar a que currentStep llegue a steps.length-1\n3. Verificar isCompleted=true en SimulationContext\n4. Verificar isPlaying=false (auto-pause al completar)',
 'Todos los pasos se ejecutan. isCompleted=true. isPlaying=false. Todos los elementos en verde.','Critica','Pendiente'],

['CP-SIM-07','Simulacion','Controles al finalizar','HU-06',
 '1. Simulacion recien completada (isCompleted=true)',
 'N/A',
 '1. Verificar que Play no funciona (play() revisa !isCompleted)\n2. Verificar boton Reiniciar (replay) funciona\n3. Presionar Reiniciar: verificar resetSimulation() ejecuta',
 'togglePlayPause() no cambia isPlaying si isCompleted=true. resetSimulation() pone currentStep=0, isPlaying=false, isCompleted=false.','Alta','Pendiente'],

['CP-SIM-08','Simulacion','Mensaje "Algoritmo completado!" con opciones','HU-07',
 '1. Simulacion recien completada',
 'N/A',
 '1. Verificar CompletionOverlay.tsx aparece\n2. Verificar 3 opciones: "Reiniciar", "Siguiente Algoritmo", "Ver Codigo"\n3. Esperar 5 segundos sin interaccion\n4. Verificar que overlay desaparece automaticamente',
 'Overlay visible con 3 botones. Desaparece a los 5s (setTimeout). No bloquea navegacion.','Alta','Pendiente'],

['CP-SIM-09','Simulacion','Timeout contra bucle infinito','HU-06',
 '1. Engine con dataset de limites',
 '15 elementos en orden inverso [15,14,...,1]',
 '1. Iniciar simulacion con peor caso\n2. Contar pasos generados por generateBubbleSortSteps()\n3. Verificar que steps.length < 10000\n4. Verificar que la app no se congela',
 'Simulacion termina. steps.length < 10000. Si excede: alerta de timeout. App responsiva.','Critica','Pendiente'],

['CP-SIM-10','Simulacion','Feedback accesible daltonicos','HU-06',
 '1. Simulacion ejecutandose',
 'N/A',
 '1. Verificar que Bar.tsx muestra iconos ademas de colores\n2. Check para completado, X para error',
 'Iconos visibles ademas de color coding en cada barra','Media','Pendiente'],

['CP-DAT-01','Datos','Datos predeterminados 8-15 elementos','HU-03',
 '1. Algoritmo seleccionado, SimulationScreen cargando',
 'Ninguno (generacion automatica)',
 '1. Navegar a SimulationScreen\n2. Verificar area de visualizacion NO vacia\n3. Contar barras: entre 8 y 15\n4. Verificar que no estan ordenadas ascendentemente',
 'Barras visibles. Cantidad entre 8 y 15. Array no ordenado. data[] en SimulationContext tiene 8-15 numeros enteros.','Critica','Pendiente'],

['CP-DAT-02','Datos','Generar nuevos datos refresca arreglo','HU-03',
 '1. Simulacion cargada con datos predeterminados',
 'Tap en "Generar nuevos datos"',
 '1. Anotar data[] actual del SimulationContext\n2. Presionar boton\n3. Verificar nuevo data[] diferente\n4. Verificar steps[] recalculados',
 'Nuevo arreglo diferente, 8-15 elementos, no ordenado. steps[] regenerados por engine.','Alta','Pendiente'],

['CP-DAT-03','Datos','Datos personalizados validos','HU-03',
 '1. Opcion datos personalizados visible',
 '[5, 2, 8, 1, 9, 3]',
 '1. Seleccionar "Datos Personalizados"\n2. Ingresar: 5, 2, 8, 1, 9, 3\n3. Confirmar\n4. Verificar data[]=[5,2,8,1,9,3] en SimulationContext',
 '6 barras con alturas proporcionales a los valores. data[] refleja input exacto.','Alta','Pendiente'],

['CP-DAT-04','Datos','Datos invalidos muestran error','Glosario',
 '1. Opcion datos personalizados visible',
 'Var A: [5,abc,3]\nVar B: [5,,3]\nVar C: (vacio)\nVar D: [5;DROP TABLE;3]',
 '1. Para cada variante: ingresar datos\n2. Presionar confirmar\n3. Verificar mensaje de error\n4. Verificar que simulacion NO inicia\n5. Verificar datos anteriores intactos',
 'Mensaje de error no tecnico. Simulacion no inicia. data[] anterior se mantiene.','Alta','Pendiente'],

['CP-CTL-01','Controles','Reiniciar simulacion','HU-06',
 '1. Simulacion en cualquier paso (currentStep > 0)',
 'Tap en boton Reiniciar (replay icon en ControlBar)',
 '1. Presionar boton replay\n2. Verificar resetSimulation() ejecutado\n3. Verificar currentStep=0, isPlaying=false, isCompleted=false',
 'Simulacion vuelve al paso 0. Estado limpio. Barras en posicion inicial.','Alta','Pendiente'],

['CP-ENG-01','Engine','Bubble Sort genera pasos correctos','HU-04',
 '1. Funcion generateBubbleSortSteps disponible',
 'Input: [5, 2, 8, 1]',
 '1. Llamar generateBubbleSortSteps([5,2,8,1])\n2. Verificar paso 0: tipoOperacion="idle", lineaPseudocodigo=2\n3. Verificar comparaciones tienen tipoOperacion="comparison"\n4. Verificar intercambios tienen tipoOperacion="swap"\n5. Verificar ultimo paso: estadoArray=[1,2,5,8] (ordenado)',
 'steps[] no vacio. Primer paso idle. Ultimo paso con array ordenado [1,2,5,8]. Cada step tiene: numeroPaso, tipoOperacion, indicesActivos, estadoArray, lineaPseudocodigo.','Critica','Pendiente'],

['CP-ENG-02','Engine','Selection Sort genera pasos correctos','HU-04',
 '1. generateSelectionSortSteps disponible',
 'Input: [5, 2, 8, 1]',
 '1. Llamar generateSelectionSortSteps([5,2,8,1])\n2. Verificar estructura de steps identica a Bubble Sort\n3. Verificar ultimo paso: estadoArray=[1,2,5,8]',
 'Array final ordenado. Steps con tipoOperacion correcto.','Critica','Pendiente'],

['CP-ENG-03','Engine','Insertion Sort genera pasos correctos','HU-04',
 '1. generateInsertionSortSteps disponible',
 'Input: [5, 2, 8, 1]',
 '1. Llamar generateInsertionSortSteps([5,2,8,1])\n2. Verificar estructura\n3. Verificar ultimo paso: estadoArray=[1,2,5,8]',
 'Array final ordenado. Steps con tipoOperacion correcto.','Critica','Pendiente'],

['CP-REN-01','Rendimiento','Simulacion >= 24 FPS','HU-04',
 '1. Dispositivo: Pixel 5a (Android 14)\n2. Expo Go instalado\n3. React DevTools Profiler conectado\n4. Bateria >50%, sin apps en foreground',
 'Bubble Sort, [15,14,13,...,1], velocidad 2.0x',
 '1. Cargar SimulationScreen con 15 elem orden inverso\n2. Abrir Profiler -> Start Recording\n3. Play a 2.0x\n4. Esperar finalizacion\n5. Exportar metricas FPS',
 'FPS promedio >= 24. FPS minimo >= 18. Sin frame drops >3 frames consecutivos.','Critica','Pendiente'],
]

for i,c in enumerate(cases):
    ws.append(c); do_row(ws, i+2, c[8])

widths = [12,14,38,10,45,30,60,50,12,12]
for i,w in enumerate(widths): ws.column_dimensions[chr(65+i)].width = w
ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = 'A2'

# HOJA 2: TRAZABILIDAD
ws2 = wb.create_sheet('Trazabilidad')
ws2.append(['Caso de Prueba','HU/CO','Modulo','Componente','Severidad'])
do_hdr(ws2)
for c in cases:
    ws2.append([c[0],c[3],c[1],c[2],c[8]]); do_row(ws2, ws2.max_row, c[8])
for i,w in enumerate([16,10,14,35,12]): ws2.column_dimensions[chr(65+i)].width = w
ws2.auto_filter.ref = ws2.dimensions; ws2.freeze_panes = 'A2'

# HOJA 3: RESUMEN
ws3 = wb.create_sheet('Resumen')
ws3.append(['Modulo','Total','Critica','Alta','Media']); do_hdr(ws3)
mods = {}
for c in cases:
    m=c[1]
    if m not in mods: mods[m]={'t':0,'Critica':0,'Alta':0,'Media':0}
    mods[m]['t']+=1; mods[m][c[8]]=mods[m].get(c[8],0)+1
for m,v in mods.items():
    ws3.append([m,v['t'],v.get('Critica',0),v.get('Alta',0),v.get('Media',0)]); do_row(ws3,ws3.max_row)
ws3.append(['TOTAL',len(cases),sum(v.get('Critica',0) for v in mods.values()),sum(v.get('Alta',0) for v in mods.values()),sum(v.get('Media',0) for v in mods.values())])
for c in ws3[ws3.max_row]: c.font=Font(bold=True); c.border=bdr
for i,w in enumerate([16,10,10,10,10]): ws3.column_dimensions[chr(65+i)].width = w

out = os.path.join(os.path.dirname(os.path.abspath(__file__)),'..','ejemplos','3.2-Casos-de-Prueba-EJEMPLO.xlsx')
wb.save(out)
print('OK: '+out)
