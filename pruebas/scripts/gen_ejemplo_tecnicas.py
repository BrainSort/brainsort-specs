import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Ejercicio 1 - Técnicas Prueba"

# Estilos
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
title_font = Font(name='Calibri', bold=True, size=14)
hfill_dark = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
hfill_light = PatternFill(start_color='2E74B5', end_color='2E74B5', fill_type='solid')
bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Título y metadata
ws.append([])
ws.append([None, None, 'DISEÑO DE PRUEBAS - TÉCNICAS DE CAJA NEGRA (BrainSort)'])
ws.cell(row=2, column=3).font = title_font
ws.append([])
ws.append([])

# Fila 5: Super Headers
ws.append([None, None, None, None, 'CLASES DE EQUIVALENCIA', None, None, None, 'VALORES AL LIMITE', None, None, None, None, None, None])
ws.merge_cells('E5:H5')
ws.merge_cells('I5:O5')
for c in ['E', 'I']:
    cell = ws[f'{c}5']
    cell.font = header_font
    cell.fill = hfill_dark
    cell.alignment = center

# Fila 6: Headers de Columnas
headers = [None, None, 'CAMPO', 'Descripción', 'Clases válidas', 'Ej. Clases Val.', 'Clases Inválidas', 'Ej. Clases Inv.', 'MIN-1', 'MIN', 'MIN+1 (opcional)', 'NORMAL', 'MAX-1 (opcional)', 'MAX', 'MAX+1']
ws.append(headers)
for col_idx in range(3, 16):
    cell = ws.cell(row=6, column=col_idx)
    cell.font = header_font
    cell.fill = hfill_light
    cell.border = bdr
    cell.alignment = center

# Datos de Casos
casos = [
    [None, None, 'Longitud de Arreglo (Datos Personalizados)', 
     'Cantidad de números ingresados por el usuario para ser ordenados. Mínimo 2, Máximo 15.', 
     '* Arreglo con longitud >= 2 y <= 15', 
     'longitud == 5\nEj: [5, 2, 8, 1, 9]', 
     '* Arreglo vacío\n* Arreglo con 1 elemento\n* Arreglo con más de 15 elementos', 
     '* longitud == 0 (Ej: [])\n* longitud == 1 (Ej: [5])\n* longitud == 16', 
     'length == 1\nEj: [5]', 
     'length == 2\nEj: [5, 2]', 
     'length == 3\nEj: [5, 2, 8]', 
     'length == 8\nEj: [5,2,8,1,9,3,7,4]', 
     'length == 14\nEj: [14 elems]', 
     'length == 15\nEj: [15 elems]', 
     'length == 16\nEj: [16 elems] (Error)'],
    
    [None, None, 'Formato de Datos (Datos Personalizados)', 
     'Cadena de texto ingresada en el input. Debe contener solo números enteros separados por comas.', 
     '* Números enteros separados por coma\n* Puede incluir números negativos', 
     '"10, -5, 3, 42"', 
     '* Contiene letras\n* Caracteres especiales distintos a coma o menos\n* Comas consecutivas\n* Puros espacios', 
     '* "10, A, 3"\n* "10; 5; 3"\n* "10,,3"\n* "   "', 
     'N/A', 'N/A', 'N/A', '"1, 2, 3"', 'N/A', 'N/A', 'N/A'],

    [None, None, 'Velocidad de Simulación (SpeedSlider)', 
     'Selector numérico de velocidad que afecta el delay de animación. Rango de 0.25x a 2.0x.', 
     '* Valor numérico entre 0.25 y 2.0', 
     '1.0', 
     '* Valor < 0.25\n* Valor > 2.0\n* Valores no numéricos (si fuesen inyectables)', 
     '* 0.1\n* 3.0\n* "rápido"', 
     '0.0\n(Fuera de rango)', 
     '0.25', 
     '0.50', 
     '1.0\n(Normal)', 
     '1.75', 
     '2.0', 
     '2.25\n(Fuera de rango)'],

    [None, None, 'Límite de Pasos de Animación (Timeout)', 
     'Mecanismo de seguridad del engine para evitar bucles infinitos. Máximo 10,000 pasos.', 
     '* Cantidad total de pasos <= 10000', 
     '150 pasos procesados', 
     '* Cantidad total de pasos > 10000', 
     '10001 pasos\n(Lanza Error de Timeout)', 
     '0\n(Array ya ordenado, 0 swaps)', 
     '1\n(Mínima iteración)', 
     '2', 
     '150\n(Uso promedio)', 
     '9999', 
     '10000\n(Justo en el límite)', 
     '10001\n(Falla por timeout)']
]

for row_data in casos:
    ws.append(row_data)
    current_row = ws.max_row
    for col_idx in range(3, 16):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.alignment = wrap
        cell.border = bdr

# Ajustar anchos de columna
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 20
for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O']:
    ws.column_dimensions[col].width = 18

# Guardar
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'ejemplos', '3.4-Clases-Equivalencia-BrainSort.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print('OK: ' + out_path)
